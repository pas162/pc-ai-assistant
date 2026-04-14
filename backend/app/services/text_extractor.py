import fitz        # PyMuPDF — for regular text extraction
import pdfplumber  # for table extraction
import docx  # python-docx
import openpyxl
import os
from typing import Callable, Optional


class ExtractionCancelled(Exception):
    """Raised when the caller signals cancellation via on_progress callback."""
    pass


def extract_text(
    file_path: str,
    file_type: str,
    on_progress: Optional[Callable[[int, int], None]] = None
) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    if file_type == "pdf":
        return _extract_from_pdf(file_path, on_progress)
    elif file_type == "docx":
        return _extract_from_docx(file_path)
    elif file_type in ("txt", "md", "csv"):
        return _extract_from_txt(file_path)
    elif file_type in ("xlsx", "xls"):
        return _extract_from_excel(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")

def _extract_from_pdf(
    file_path: str,
    on_progress: Optional[Callable[[int, int], None]] = None
) -> str:
    page_parts = []
    print(f"[Extractor] Opening PDF: {file_path}")

    with fitz.open(file_path) as fitz_pdf:
        total_pages = len(fitz_pdf)
        print(f"[Extractor] Total pages: {total_pages}")

        plumber_pdf = pdfplumber.open(file_path)
        try:
            for page_num in range(total_pages):
                fitz_page    = fitz_pdf[page_num]
                plumber_page = plumber_pdf.pages[page_num]
                page_label   = f"[Page {page_num + 1} of {total_pages}]"
                parts        = [page_label]

                # ── Fast check: does this page likely have a table? ────────
                # Use PyMuPDF first — it's instant
                # Only call pdfplumber (slow) if we see table-like structure
                plain_text = fitz_page.get_text().strip()
                has_table_hint = len(fitz_page.get_drawings()) > 5  # lines/rects = table borders

                table_bboxes = []
                if has_table_hint:
                    try:
                        tables = plumber_page.extract_tables()
                        if tables:
                            print(f"[Extractor] Page {page_num + 1}: found {len(tables)} tables")
                        for table in tables:
                            if not table:
                                continue
                            md_rows = []
                            for i, row in enumerate(table):
                                clean_row = [cell.strip() if cell else "" for cell in row]
                                md_rows.append("| " + " | ".join(clean_row) + " |")
                                if i == 0:
                                    md_rows.append("|" + "|".join(["---"] * len(clean_row)) + "|")
                            parts.append("\n".join(md_rows))
                            bbox = plumber_page.find_tables()[len(table_bboxes)].bbox
                            table_bboxes.append(bbox)
                    except Exception as e:
                        print(f"[Extractor] Table extraction failed on page {page_num + 1}: {e}")
                        table_bboxes = []

                # ── Extract regular text via PyMuPDF ──────────────────────
                if table_bboxes:
                    text_parts = []
                    for block in fitz_page.get_text("blocks"):
                        bx0, by0, bx1, by1, block_text = block[:5]
                        block_text = block_text.strip()
                        if not block_text:
                            continue
                        in_table = any(
                            bx0 < tx1 and bx1 > tx0 and by0 < ty1 and by1 > ty0
                            for (tx0, ty0, tx1, ty1) in table_bboxes
                        )
                        if not in_table:
                            text_parts.append(block_text)
                    if text_parts:
                        parts.append("\n".join(text_parts))
                else:
                    # No tables — use fast plain text only
                    if plain_text:
                        parts.append(plain_text)

                if len(parts) > 1:
                    page_parts.append("\n\n".join(parts))

                # ── Report progress after every page ──────────────────────
                if on_progress:
                    on_progress(page_num + 1, total_pages)
                    # ↑ on_progress raises ExtractionCancelled if doc deleted
                    # The exception bubbles up through the for loop,
                    # hits the finally block (closes pdfplumber), then
                    # propagates to document_processor.py

        finally:
            plumber_pdf.close()  # ← always runs, even on ExtractionCancelled
            print(f"[Extractor] Extraction complete: {len(page_parts)} pages")

    return "\n\n".join(page_parts)


def _extract_from_docx(file_path: str) -> str:
    """
    Extract text from a DOCX file using python-docx.
    A DOCX file is actually a ZIP file containing XML.
    python-docx handles all that complexity for us.
    """
    doc = docx.Document(file_path)

    # Each paragraph is like a line or block of text in the document
    paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]

    return "\n\n".join(paragraphs)


def _extract_from_txt(file_path: str) -> str:
    """
    Extract text from a plain text file.
    Simple file read — no special library needed.
    """
    # Try UTF-8 first, fall back to latin-1 if it fails
    # (some files have special characters that break UTF-8)
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="latin-1") as f:
            return f.read()


def _extract_from_excel(file_path: str) -> str:
    """
    Extract text from an Excel file using openpyxl.

    Strategy:
    - Loop through every sheet
    - Convert each sheet's data to a markdown table
    - Label each sheet clearly so the LLM knows which sheet data came from

    Example output:
        [Sheet: Product List]
        | ID | Name  | Price |
        |----|-------|-------|
        | 1  | ItemA | 100   |
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    # data_only=True → reads cell VALUES not formulas
    # e.g. cell with =SUM(A1:A10) returns 550, not the formula string

    sheet_parts = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Skip completely empty sheets
        if sheet.max_row == 0 or sheet.max_column == 0:
            continue

        # Read all rows, convert None cells to empty string
        rows = []
        for row in sheet.iter_rows(values_only=True):
            # Skip rows that are entirely empty
            if all(cell is None for cell in row):
                continue
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            rows.append(clean_row)

        if not rows:
            continue

        # Convert to markdown table
        # First row = header
        md_rows = []
        md_rows.append("| " + " | ".join(rows[0]) + " |")
        md_rows.append("|" + "|".join(["---"] * len(rows[0])) + "|")

        for row in rows[1:]:
            # Pad row if it has fewer columns than header
            while len(row) < len(rows[0]):
                row.append("")
            md_rows.append("| " + " | ".join(row) + " |")

        sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(md_rows)
        sheet_parts.append(sheet_text)

        print(f"[Extractor] Excel sheet '{sheet_name}': {len(rows)} rows extracted")

    if not sheet_parts:
        return ""

    return "\n\n".join(sheet_parts)
